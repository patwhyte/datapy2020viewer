#-------------------------------------------------------------------------------
# Name:     pyDataViewer
# Purpose:  This program enables a user to read an excel file and export it to
#           a db.
# Author:   Pat White
# Created:  03/12/2020 for CSU | CPSC4205 | Senior Project
#-------------------------------------------------------------------------------

#importing libs for use in prgram
import openpyxl
import os
import time
import sys
import pyfiglet
import sqlite3

'''
classes and functions in alphabetical order
'''

'''class ExcelRecord used for creating objects from the excel and transfer
the records to a databse'''
class ExcelRecord:
    def __init__(self, segment, country, product, disBand, uSold, manPrice,
        salePrice,saleGross, discount, sale, cogs, profit, date, monNum,
        monName, year):
        self.colA = segment
        self.colB = country
        self.colC = product
        self.colD = disBand
        self.colE = uSold
        self.colF = manPrice
        self.colG = salePrice
        self.colH = saleGross
        self.colI = discount
        self.colJ = sale
        self.colK = cogs
        self.colL = profit
        self.colM = date
        self.colN = monNum
        self.colO = monName
        self.colP = year


''' func addExcelRow no args & return. Used to add new row to the excel file'''
def addExcelRow():
    print('>>>Adding a new record<<<')
    newEntryRow = sheet.max_row + 1
    print('New record will be added in row ', newEntryRow, '.')
    newRecord = ExcelRecord("", "", "","", "","", "","", "","", "","", "","",\
        "", "")
    addRow = False
    rowModify = False

    #loop to user in function
    while not addRow:
        userChoice = input(str('Do you want to add a record to the excel (1)'\
            'or cancel the add (99)?'))

        #checking if user wants to add a row
        if userChoice == '99':
            print('You canceled the add. No changes made to the excel.')
            return

        #if user wants to add new row create new object with provided data
        elif userChoice == '1':
            receiveAddRecord(newRecord)     #call to add new object
            print()
            printAddedRecord(newRecord)     #call to print the new object
            print()

            #giving user the choice to make modifications to the input given
            while not rowModify:
                userChoice = input(str('Would you like to make changes to'\
                    ' the fields (1) or keep the data (99)?'))

                if userChoice == '1':
                    continueToSave = False

                    #user provides selection and then can edit a field
                    while not continueToSave:
                        userChoice = input(str('Please enter the edit number'\
                            ' (1-16) to change the field.'))
                        editExcelField(newRecord, userChoice)
                        correctInput = False

                        while not correctInput:
                            userChoice = input(str('Would you like to change'\
                                ' another field (1) or continue to save'\
                                ' process (99)?'))

                            if userChoice == '99':
                                print('Continuing to save process.')
                                correctInput = True
                                continueToSave = True
                                rowModify = True
                                addRow =True

                            elif userChoice == '1':
                                correctInput = True

                            else:
                                print('Error - Please enter a number 1-16 or'\
                                    ' 99 to continue without changes.')

                elif userChoice == '99':
                    rowModify = True
                    addRow = True

                else:
                    print('Error - Please enter a 1 to make changes or 99 to'\
                    ' continue without changes.')

        else:
            print('Error - Please enter a 1 to add a record or 99 to cancel.')

    #write the user created object to the excel file
    populateExcel(newRecord, newEntryRow)       #call to write to excel
    wb.save(defaultFile)                        #save the excel
    del newRecord                               #del object for next use


'''function editExcelField takes record (the empty object addExcelRow) and
number (indicated number by user to modify from addExcelRow). No return.'''
def editExcelField(record, number):
    newRecord = record
    #checking the number and changing the corresponding property
    if number == '1':
        print('Current SEGMENT is: ', newRecord.colA)
        userChoice = input(str('Please enter a new SEGMENT: '))
        newRecord.colA = userChoice
        print('SEGMENT field changed to ', newRecord.colA, '.')

    elif number == '2':
        print('Current COUNTRY is: ', newRecord.colB)
        userChoice = input(str('Please enter a new COUNTRY: '))
        newRecord.colB = userChoice
        print('COUNTRY field changed to ', newRecord.colB, '.')

    elif number == '3':
        print('Current PRODUCT is: ', newRecord.colC)
        userChoice = input(str('Please enter a new PRODUCT: '))
        newRecord.colC = userChoice
        print('PRODUCT field changed to ', newRecord.colC, '.')

    elif number == '4':
        print('Current DISCOUNT BAND is: ', newRecord.colD)
        userChoice = input(str('Please enter a new DISCOUNT BAND: '))
        newRecord.colD = userChoice
        print('DISCOUNT BAND field changed to ', newRecord.colD, '.')

    elif number == '5':
        print('Current UNITS SOLD is: ', newRecord.colE)
        userChoice = input(str('Please enter a new UNITS SOLD: '))
        newRecord.colE = userChoice
        print('UNITS SOLD field changed to ', newRecord.colE, '.')

    elif number == '6':
        print('Current MANUFACTURING PRICE is: ', newRecord.colF)
        userChoice = input(str('Please enter a new MANUFACTURING PRICE: '))
        newRecord.colF = userChoice
        print('MANUFACTURING PRICE field changed to ', newRecord.colF, '.')

    elif number == '7':
        print('Current SALES PRICE is: ', newRecord.colG)
        userChoice = input(str('Please enter a new SALES PRICE: '))
        newRecord.colG = userChoice
        print('SALES PRICE field changed to ', newRecord.colG, '.')

    elif number == '8':
        print('Current GROSS SALES is: ', newRecord.colH)
        userChoice = input(str('Please enter a new GROSS SALES: '))
        newRecord.colH = userChoice
        print('GROSS SALES field changed to ', newRecord.colH, '.')

    elif number == '9':
        print('Current SALES is: ', newRecord.colI)
        userChoice = input(str('Please enter a new SALES: '))
        newRecord.colI = userChoice
        print('SALES field changed to ', newRecord.colI, '.')

    elif number == '10':
        print('Current DISCOUNTS is: ', newRecord.colJ)
        userChoice = input(str('Please enter a new DISCOUNTS: '))
        newRecord.colJ = userChoice
        print('DISCOUNTS field changed to ', newRecord.colJ, '.')

    elif number == '11':
        print('Current COGS is: ', newRecord.colK)
        userChoice = input(str('Please enter a new COGS: '))
        newRecord.colK = userChoice
        print('COGS field changed to ', newRecord.colK, '.')

    elif number == '12':
        print('Current PROFIT is: ', newRecord.colL)
        userChoice = input(str('Please enter a new PROFIT: '))
        newRecord.colL = userChoice
        print('PROFIT field changed to ', newRecord.colL, '.')

    elif number == '13':
        print('Current DATE is: ', newRecord.colM)
        userChoice = input(str('Please enter a new DATE: '))
        newRecord.colM = userChoice
        print('DATE field changed to ', newRecord.colM, '.')

    elif number == '14':
        print('Current MONTH NUMBER is: ', newRecord.colN)
        userChoice = input(str('Please enter a new MONTH NUMBER: '))
        newRecord.colN = userChoice
        print('MONTH NUMBER field changed to ', newRecord.colN, '.')

    elif number == '15':
        print('Current MONTH NAME is: ', newRecord.colO)
        userChoice = input(str('Please enter a new MONTH NAME: '))
        newRecord.colO = userChoice
        print('MONTH NAME field changed to ', newRecord.colO, '.')

    elif number == '16':
        print('Current YEAR is: ', newRecord.colP)
        userChoice = input(str('Please enter a new YEAR: '))
        newRecord.colP = userChoice
        print('YEAR field changed to ', newRecord.colP, '.')


'''func exortToDb takes and returns oldDbPath. Used to export data to a db.'''
def exortToDb(oldDbPath):
    print('The current db path is: ', oldDbPath)
    changeDbPath = False
    pathGood = False

    #check with user if path is good
    while not changeDbPath:
        userChoice = input(str('Would you like to change the path (1) or '\
            'keep current (99)?'))

        if userChoice == '99':
            print('No change to path. Path remains... ', oldDbPath)
            changeDbPath = True

        elif userChoice == '1':

            #provide user option to change path
            while not pathGood:
                newPath = input(str('Please enter a new path: '))
                #checking the path
                try:
                    conn = sqlite3.connect(newPath)
                    conn.close()
                    pathGood = True
                    print('The path is now set to: ', newPath)
                    oldDbPath = newPath
                    changeDbPath = True

                except IOError:
                    print('You entered: ', newPath)
                    print('The system path could not be found.')
                    exitPath = False

                    while not exitPath:
                        userChoice = input(str('Do you want to try again (1) '\
                        'or keep old file path (99)?'))

                        if userChoice == '1' or '99':

                            if userChoice == '99':
                                exitPath = True
                                pathGood = True
                                changeDbPath =True
                                print('Keeping old file path...', oldDbPath)

                            elif userChoice == '1':
                                exitPath = True

                            else:
                                print('Warning - Invalid input.')

    #connecting to db and transfering data
    connectionDb = sqlite3.connect(oldDbPath)
    print('Connecting to database at: ', dbComplete)

    a = 'Channel Partners'
    b = 'Enterprise'
    c = 'Government'
    d = 'Midmarket'
    e = 'Small Business'
    rowCounter = 2              #is 2 because excel row 1 is column name
    rowLimit = False

    #loop to iterate through the data
    while not rowLimit:
        record = sheet.cell(row = rowCounter, column = 1).value

        #check to make sure the row contains a specified segment vars a-e
        if record == a or record == b or record == c or record == d or\
            record == e:
            segment = str(sheet.cell(row = rowCounter, column = 1).value)
            country = str(sheet.cell(row = rowCounter, column = 2).value)
            product = str(sheet.cell(row = rowCounter, column = 3).value)
            discountBand = str(sheet.cell(row = rowCounter, column = 4).value)
            unitsSold = str(sheet.cell(row = rowCounter, column = 5).value)
            manufacturingPrice = str(sheet.cell(row = rowCounter, column = 6)\
                .value)
            salePrice = str(sheet.cell(row = rowCounter, column = 7).value)
            grossSales = str(sheet.cell(row = rowCounter, column = 8).value)
            discounts = str(sheet.cell(row = rowCounter, column = 9).value)
            sales = str(sheet.cell(row = rowCounter, column = 10).value)
            cogs = str(sheet.cell(row = rowCounter, column = 11).value)
            profit = str(sheet.cell(row = rowCounter, column = 12).value)
            date = str(sheet.cell(row = rowCounter, column = 13).value)
            monthNumber = str(sheet.cell(row = rowCounter, column = 14).value)
            monthName = str(sheet.cell(row = rowCounter, column = 15).value)
            year = str(sheet.cell(row = rowCounter, column = 16).value)

            #sql for db, ? correspond to data in var tableValues
            sqlStatement = 'INSERT INTO SENIORPROJECT (Segment, Country,\
                Product, DiscountBand, UnitsSold, ManufacturingPrice,\
                SalePrice, GrossSales, Discounts, Sales, COGS, Profit, Date,\
                MonthNumber, MonthName, Year) VALUES (?, ?, ?, ?, ?, ?, ?, ?,\
                ?, ?, ?, ?, ?, ?, ?, ?)'

            tableValues = (segment, country, product, discountBand, unitsSold,
                manufacturingPrice, salePrice, grossSales, discounts, sales,\
                cogs, profit, date, monthNumber, monthName, year)

            #connecting and transfering
            with connectionDb:
                connectionDb.execute(sqlStatement, tableValues)
                connectionDb.commit()                   #saving to db

            rowCounter += 1

        else:
            rowLimit = True

    rowCounter -= 2
    print('Transfer in progress...please standby by.')
    time.sleep(3)
    print('Transfer complete.', rowCounter, 'records were exported.')
    connectionDb.close()
    return oldDbPath


'''func mainDisplay takes no args/no return. Used to display a main menu.'''
def mainDisplay():
    print()
    astLine = '*' * 76
    print(astLine)
    print(astLine)
    #ascii art using pyfiglet
    menuHeader = pyfiglet.figlet_format("  MAIN MENU", font = "bubble")
    print(menuHeader)
    print('    Please choose an option below:')
    print(' ------------------------------------')
    print('  Sales Information ------> Enter 1')
    print('  Product Information ----> Enter 2')
    print('  Add Excel Entry --------> Enter 3')
    print('  Export to DB -----------> Enter 4')
    print('  Change path/file/sheet -> Enter 5')
    print('  Option Details ---------> Enter 77')
    print('  Exit the Program -------> Enter 99')
    print()
    print(astLine)
    print(astLine)


'''func optionDetails noarg/no return. used to display a short description
of what each menu item does'''
def optionDetails():
    astLine = '*' * 76
    print(astLine)
    print(astLine)

    print('Sales Information --------> This will display the totals for\n'\
        '\t\t\t\t\t\t\tsales, cost of sales, and profits for each of the\n '\
        '\t\t\t\t\t\t\tsegments with final totals.')
    print('Product Information ------> This will display each product with\n'\
        '\t\t\t\t\t\t\tnumber of regions reporting, the units they sold,\n'\
        '\t\t\t\t\t\t\tand their profit.')
    print('Add Excel Entry ----------> This enables the user to add\n'\
        '\t\t\t\t\t\t\tadditional rows to the excel file. When selected the\n'\
        '\t\t\t\t\t\t\tuser must provide input for each column on the sheet.')
    print('Export to DB -------------> The option enables the user to export\n'\
        '\t\t\t\t\t\t\tthe excel data to a data base. User can specify which\n'\
        '\t\t\t\t\t\t\tdata base to use.')
    print('Change path/file/sheet ---> Here the user can change the location\n'\
        '\t\t\t\t\t\t\tof the file, the name, and the sheet\n'\
        '\t\t\t\t\t\t\tto be read.')
    print('Option Details -----------> This menu here.')
    print('Exit the Program ---------> Closes the connection to the file\n'\
        '\t\t\t\t\t\t\tand program.')

    print(astLine)
    print(astLine)
    hold = input("Press any key to continue.")
    print()


'''func populateExcel takes record (the user created object from add
addExcelRow) and row (the row that will be written to), no return. Used to
write data to excel)'''
def populateExcel(record, row):
    newRecord = record
    a = 'A' + str(row)
    sheet[a] = newRecord.colA
    b = 'B' + str(row)
    sheet[b] = newRecord.colB
    c = 'C' + str(row)
    sheet[c] = newRecord.colC
    d = 'D' + str(row)
    sheet[d] = newRecord.colD
    e = 'E' + str(row)
    sheet[e] = newRecord.colE
    f = 'F' + str(row)
    sheet[f] = newRecord.colF
    g = 'G' + str(row)
    sheet[g] = newRecord.colG
    h = 'H' + str(row)
    sheet[h] = newRecord.colH
    i = 'I' + str(row)
    sheet[i] = newRecord.colI
    j = 'J' + str(row)
    sheet[j] = newRecord.colJ
    k = 'K' + str(row)
    sheet[k] = newRecord.colK
    l = 'L' + str(row)
    sheet[l] = newRecord.colL
    m = 'M' + str(row)
    sheet[m] = newRecord.colM
    n = 'N' + str(row)
    sheet[n] = newRecord.colN
    o = 'O' + str(row)
    sheet[o] = newRecord.colO
    p = 'P' + str(row)
    sheet[p] = newRecord.colP


'''func printAddedRecord takes record (the created object) no return. Used
to display the input of the record to the user prior to saving data to excel'''
def printAddedRecord(record):
    newRecord = record
    #printing the coulmn name, the input and a value for identification
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('Column', 'Input', 'Modify Value'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('SEGMENT', newRecord.colA, '1'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('COUNTRY', newRecord.colB, '2'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('PRODUCT', newRecord.colC, '3'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('DISCOUNT BAND', newRecord.colD, '4'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('UNITS SOLD', newRecord.colE, '5'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('MANUFACTURING PRICE', newRecord.colF, '6'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('SALES PRICE', newRecord.colG, '7'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('GROSS SALES', newRecord.colH, '8'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('DISCOUNTS', newRecord.colI, '9'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('SALES', newRecord.colJ, '10'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('COGS', newRecord.colK, '11'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('PROFIT', newRecord.colL, '12'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('DATE', newRecord.colM, '13'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('MONTH NUMBER', newRecord.colN, '14'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('MONTH NAME', newRecord.colP, '15'))
    print('{:' '<24s} {:' '<24s} {:' '>14s}'.format\
        ('YEAR', newRecord.colP, '16'))


'''func productInformation takes no args, no returns. used to display info
to user based on products'''
def productInformation():
    print('>>>>>>>>Retreiving data<<<<<<<<<<<')
    print()
    astLine = '*' * 76
    rowLimit = False
    rowCounter = 2          #start at 2 because of excel header

    #columns we want to get data from
    numCol = 3
    unitsCol = 5
    profitCol = 12

    amaCount = 0
    carCount = 0
    monCount = 0
    pasCount = 0
    velCount = 0
    vttCount = 0

    amaUnits = 0
    carUnits = 0
    monUnits = 0
    pasUnits = 0
    velUnits = 0
    vttUnits = 0

    amaProfit = 0
    carProfit = 0
    monProfit = 0
    pasProfit = 0
    velProfit = 0
    vttProfit = 0

    #loop to iterate through the sheet
    while not rowLimit:

        #getting the data from matching row (counter based) for each segment
        if str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Amarilla':
            amaCount += 1
            amaUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            amaProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Carretera':
            carCount += 1
            carUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            carProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Montana':
            monCount += 1
            monUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            monProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Paseo':
            pasCount += 1
            pasUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            pasProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) == 'Velo':
            velCount += 1
            velUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            velProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) == 'VTT':
            vttCount += 1
            vttUnits += int(sheet.cell(row = rowCounter, column = unitsCol)\
            .value)
            vttProfit += int(sheet.cell(row = rowCounter, column = profitCol)\
            .value)
            rowCounter += 1

        else:
            rowLimit = True
            rowCounter -= 2     #taking away 2 because started with 2

    #getting totals
    regionTotal = amaCount  + carCount  + monCount  + pasCount  + velCount\
        + vttCount
    unitTotal   = amaUnits  + carUnits  + monUnits  + pasUnits  + velUnits\
        + vttUnits
    profitTotal = amaProfit + carProfit + monProfit + pasProfit + velProfit\
        + vttProfit

    #printing out the results
    print(astLine)
    print(astLine)

    productHeader = pyfiglet.figlet_format("          PRODUCT INFORMATION",\
        font = "digital")
    print(productHeader)
    print('There are ' + str(rowCounter) + ' regions'\
        ' reporting.')
    print()
    print('-' * 58)
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>18s}'.format('Product', \
        'Regions', 'Units Sold', 'Profit'))
    print('-' * 54)
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Amarilla', str(amaCount), str(amaUnits), '$', str(amaProfit)))
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Carretera', str(carCount), str(carUnits), '$', str(carProfit)))
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Montana', str(monCount), str(monUnits), '$', str(monProfit)))
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Paseo', str(pasCount), str(pasUnits), '$', str(pasProfit)))
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Velo', str(velCount), str(velUnits), '$', str(velProfit)))
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'VTT', str(vttCount), str(vttUnits), '$', str(vttProfit)))
    print('=' * 58)
    print('{:' '<14s} {:' '>8s} {:' '>14s} {:' '>5s} {:_>12s}'.format(\
        'Final Totals', str(regionTotal), str(unitTotal), '$', str(\
        profitTotal)))

    print(astLine)
    print(astLine)
    print('<<<Press any key to continue>>>')
    hold = input()


'''func receiveAddRecord takes record (the new object) no return. Used to
populate the object properties based on user input.'''
def receiveAddRecord(record):
    newRecord = record
    print('Please provide data for each column when promted.')
    userChoice = input(str('Please enter the SEGMENT: '))
    newRecord.colA = userChoice
    userChoice = input(str('Please enter the COUNTRY: '))
    newRecord.colB = userChoice
    userChoice = input(str('Please enter the PRODUCT: '))
    newRecord.colC = userChoice
    userChoice = input(str('Please enter the DISCOUNT BAND: '))
    newRecord.colD = userChoice
    userChoice = input(str('Please enter the UNITS SOLD: '))
    newRecord.colE = userChoice
    userChoice = input(str('Please enter the MANUFACTURING PRICE: '))
    newRecord.colF = userChoice
    userChoice = input(str('Please enter the SALES PRICE: '))
    newRecord.colG = userChoice
    userChoice = input(str('Please enter the GROSS SALES: '))
    newRecord.colH = userChoice
    userChoice = input(str('Please enter the DISCOUNTS: '))
    newRecord.colI = userChoice
    userChoice = input(str('Please enter the SALES: '))
    newRecord.colJ = userChoice
    userChoice = input(str('Please enter the COGS: '))
    newRecord.colK = userChoice
    userChoice = input(str('Please enter the PROFIT: '))
    newRecord.colL = userChoice
    userChoice = input(str('Please enter the DATE: '))
    newRecord.colM = userChoice
    userChoice = input(str('Please enter the MONTH NUMBER: '))
    newRecord.colN = userChoice
    userChoice = input(str('Please enter the MONTH NAME: '))
    newRecord.colO = userChoice
    userChoice = input(str('Please enter the YEAR: '))
    newRecord.colP = userChoice

'''func salesInfo takes no args/no return. Used to display sales information
to the user'''
def salesInfo():
    print('>>>>>>>>Retreiving data<<<<<<<<<<<')
    print()

    astLine = '*' * 76
    rowLimit = False
    rowCounter = 2
    #columns that we want to display and get data from
    numCol = 1
    salesCol = 10
    cogsCol = 11
    profitCol = 12

    chanCount = 0
    entCount = 0
    govCount = 0
    midCount = 0
    smallCount = 0

    chanSalesTot = 0
    chanCogsTot = 0
    chanProfitTot = 0
    entSalesTot = 0
    entCogsTot = 0
    entProfitTot = 0
    govSalesTot = 0
    govCogsTot = 0
    govProfitTot = 0
    midSalesTot = 0
    midCogsTot = 0
    midProfitTot = 0
    smallSalesTot = 0
    smallCogsTot = 0
    smallProfitTot = 0

    #loop to iterate through the sheet based on segement
    while not rowLimit:

        if str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Channel Partners':
            chanCount += 1
            chanSalesTot += int(sheet.cell(row = rowCounter, column =\
            salesCol).value)
            chanCogsTot += int(sheet.cell(row = rowCounter, column =\
            cogsCol).value)
            chanProfitTot += int(sheet.cell(row = rowCounter, column =\
            profitCol).value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
            'Enterprise':
            entCount += 1
            entSalesTot += int(sheet.cell(row = rowCounter, column =\
            salesCol).value)
            entCogsTot += int(sheet.cell(row = rowCounter, column =\
            cogsCol).value)
            entProfitTot += int(sheet.cell(row = rowCounter, column =\
            profitCol).value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
         'Government':
            govCount += 1
            govSalesTot += int(sheet.cell(row = rowCounter, column =\
            salesCol).value)
            govCogsTot += int(sheet.cell(row = rowCounter, column =\
            cogsCol).value)
            govProfitTot += int(sheet.cell(row = rowCounter, column =\
            profitCol).value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
         'Midmarket':
            midCount += 1
            midSalesTot += int(sheet.cell(row = rowCounter, column =\
            salesCol).value)
            midCogsTot += int(sheet.cell(row = rowCounter, column =\
            cogsCol).value)
            midProfitTot += int(sheet.cell(row = rowCounter, column =\
            profitCol).value)
            rowCounter += 1

        elif str(sheet.cell(row = rowCounter, column = numCol).value) ==\
         'Small Business':
            smallCount += 1
            smallSalesTot += int(sheet.cell(row = rowCounter, column =\
            salesCol).value)
            smallCogsTot += int(sheet.cell(row = rowCounter, column =\
            cogsCol).value)
            smallProfitTot += int(sheet.cell(row = rowCounter, column =\
            profitCol).value)
            rowCounter += 1

        else:
            rowLimit = True
            rowCounter -= 2

    #getting totals
    salesTotal =  chanSalesTot  + entSalesTot  + govSalesTot  + midSalesTot\
        + smallCount
    cogsTotal =   chanCogsTot   + entCogsTot   + govCogsTot   + midCogsTot\
        + smallCogsTot
    profitTotal = chanProfitTot + entProfitTot + govProfitTot + midProfitTot\
        + smallProfitTot

    #setting up the print for displatying the results
    print(astLine)
    print(astLine)
    saleHeader = pyfiglet.figlet_format("                SALES INFORMATION",\
        font = "digital")
    print(saleHeader)
    print('There are ' + str(rowCounter)+ ' records in the excel file.')
    print()
    print('-' * 75)
    print('{:' '<24s} {:' '>13s} {:' '>17s} {:' '>17s}'.format(\
        'Segment', 'Sales', ' Cost of Sales', 'Profit'))
    print('-' * 74)
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Channel Partners', '$', str(chanSalesTot), '$'\
        , str(chanCogsTot), '$', str(chanProfitTot)))
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Enterprise', '$', str(entSalesTot), '$'\
        , str(entCogsTot), '$', str(entProfitTot)))
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Government', '$', str(govSalesTot), '$'\
        , str(govCogsTot), '$', str(govProfitTot)))
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Midmarket', '$', str(midSalesTot), '$'\
        , str(midCogsTot), '$', str(midProfitTot)))
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Small Business', '$', str(smallSalesTot), '$'\
        , str(smallCogsTot), '$', str(smallProfitTot)))
    print('=' * 75)
    print('{:' '<20s} {:' '>3s} {:_>13s} {:' '>3s} {:_>13s} {:' '>3s} '\
        '{:_>13s}'.format('Final Totals', '$', str(salesTotal), '$'\
        , str(cogsTotal), '$', str(profitTotal)))

    print(astLine)
    print(astLine)
    print('<<<Press any key to continue>>>')
    hold = input()


'''function setPath takes oldPath (current file path), no return. Used to modify
the current path'''
def setPath(oldPath):
    print('The current path is: ', oldPath)
    changePath = False
    pathGood = False

    #loop to keep user until a good path is provided
    while not changePath:
        userChoice = input(str('Would you like to change the path (1) or '\
            'keep current (99)?'))

        if userChoice == '99':
            print('No change to path. Path remains... ', oldPath)
            return oldPath

        elif userChoice == '1':

            #checking the user path
            while not pathGood:
                newPath = input(str('Please enter a new path: '))

                try:
                    os.chdir(newPath)
                    pathGood = True
                    print('The path is now set to: ', newPath)
                    return newPath

                except IOError:
                    print('You entered: ', newPath)
                    print('The system path could not be found.')

                    exitPath = False

                    while not exitPath:
                        userChoice = input(str('Do you want to try again (1) '\
                        'or keep old file path (99)?'))

                        if userChoice == '1' or '99':

                            if userChoice == '99':
                                exitPath = True
                                pathGood = True
                                print('Keeping old file path...', oldPath)
                                return oldPath

                            elif userChoice == '1':
                                exitPath = True

                            else:
                                print('Warning - Invalid input.')

        else:
            print('Warning - Invalid input.')


'''func setFile takes oldFile (current file location), no return. Used '''
def setFile(oldFile):
    lastFiveR = len(oldFile) - 5
    msgFile = oldFile[:lastFiveR]
    print('The current file is: ', msgFile)
    changeFile = False
    fileGood = False

    #keeping user in loop til he exits or provides good file
    while not changeFile:
        userChoice = input(str('Would you like to change the file (1) or '\
            'keep current (99)?'))

        if userChoice == '99':
            print('No change to file. File remains... ', oldFile)
            return oldFile

        elif userChoice == '1':

            #testing if provided file is good
            while not fileGood:
                newFile = input(str('Please enter a new file name: '))
                newFile = newFile + '.xlsx'

                try:
                    wb = openpyxl.load_workbook(newFile, read_only=False)
                    fileGood = True
                    print('The file name is now: ', newFile)
                    return newFile

                except IOError:
                    print('You entered: ', newFile)
                    print('The file could not be found.')
                    exitFile = False

                    while not exitFile:
                        userChoice = input(str('Do you want to try again (1)'\
                            'or keep old file name (99)?'))

                        if userChoice == '1' or '99':

                            if userChoice == '99':
                                exitFile = True
                                fileGood = True
                                print('Keeping old file name...', oldFile)
                                return oldFile

                            elif userChoice == '1':
                                exitFile = True

                            else:
                                print('Warning - Invalid input.')
        else:
            print('Warning - Invalid input.')


'''func setSheet takes oldSheet (cur sheetname) and wb (the workbook as in
filename), no returns. Used to change the sheet the data is located on'''
def setSheet(oldSheet, wb):
    print('Getting sheet names of the workbook. The following sheets are '\
    'present in the workbook:')
    i = 0

    for names in wb.sheetnames:
        print(wb.sheetnames[i])
        i += 1

    print('The current sheet is:', oldSheet)
    changeSheet = False
    sheetGood = False

    #keep user in loop until exit or proper sheet name is provided
    while not changeSheet:
        userChoice = input(str('Would you like to change the sheet (1) or '\
            'keep current (99)?'))

        if userChoice == '99':
            print('No change to sheet. Sheet remains... ', oldSheet)
            return oldSheet

        elif userChoice == '1':

            #testing sheet name
            while not sheetGood:
                newSheet = input(str('Please enter a new sheet name: '))
                i = 0

                for names in wb.sheetnames:

                    if wb.sheetnames[i] == newSheet:
                        print('The sheet name is now: ', newSheet)
                        return newSheet

                    else:
                        i += 1

                print('You entered: ', newSheet)
                print('The sheet could not be found.')
                exitSheet = False

                while not exitSheet:
                    userChoice = input(str('Do you want to try again (1)'\
                        'or keep old sheet name (99)?'))

                    if userChoice == '1' or '99':

                        if userChoice == '99':
                            exitSheet = True
                            sheetGood = True
                            print('Keeping old sheet name...', oldSheet)
                            return oldSheet

                        elif userChoice == '1':
                            exitSheet = True

                        else:
                            print('Warning - Invalid input.')

        else:
            print('Warning - Invalid input.')


''' func startPage takes no args, no return. Used for one time display when
program opens'''
def startPage():
    astLine = '*' * 76
    print(astLine)
    print(astLine)
    appHeader = pyfiglet.figlet_format("  pyDataViewer", font = "slant"  )
    print(appHeader)
    print('Read | Explore | Store Excel >>>>> to a DB                         '\
        ' PCW2020')
    print(astLine)
    print(astLine)


'''
main starts here...make changes to defaults below-----------
'''
#setting default path, file, and sheet
defaultPath = 'c:\\users\PatWhite\Desktop'
defaultFile = 'Book1.xlsx'
os.chdir(defaultPath)           #using os lib to open the path
wb = openpyxl.load_workbook(defaultFile, read_only=False) #open workbook
defaultSheet = str(wb.sheetnames[0])        #assign the 1st sheet as default

closeApp = False
validInteger = False

#calling functions to display inital default path, file, sheet to user
startPage()
print()
print('>>>>>>>>Setting up path, file name, and sheet<<<<<<<<<<<.')
print()
print('>>>Path to excel file<<<')
defaultPath = setPath(defaultPath)
print()
print('>>>Excel file name<<<')
defaultFile = setFile(defaultFile)
print()
print('>>>Excel sheet name<<<')
wb = openpyxl.load_workbook(defaultFile, read_only=False)
defaultSheet = setSheet(defaultSheet, wb)
sheet = wb[defaultSheet]

#setting the default db location...change default here
dbFileLocation = 'c:\\users\PatWhite\Desktop'
dbFileName = 'termProject.sqlite'
dbComplete = str(dbFileLocation + '\\' + dbFileName)

#main loop until user option is exit.
while not closeApp:
    while not validInteger:
        mainDisplay()

        userChoice = input('Which task do you want to do? ')

        #checking user input for selected option
        if userChoice.isdigit():
            userChoice = int(userChoice)

            #checking if the value is a choice in the menu
            if (userChoice > 0 and userChoice < 6) or userChoice == 99\
             or userChoice == 77:
                validInteger = True

                #sales info option
                if userChoice == 1:
                    validInteger = False
                    salesInfo()

                #product info option
                elif userChoice == 2:
                    validInteger = False
                    productInformation()

                #add row option
                elif userChoice == 3:
                    validInteger = False
                    addExcelRow()

                #transfer to db option
                elif userChoice == 4:
                    validInteger = False
                    dbComplete = exortToDb(dbComplete)

                #change defaults option
                elif userChoice == 5:
                    validInteger = False
                    defaultPath = setPath(defaultPath)
                    defaultFile = setFile(defaultFile)
                    defaultSheet = setSheet(defaultSheet, wb)
                    sheet = wb[defaultSheet]

                #display details option
                elif userChoice == 77:
                    validInteger = False
                    optionDetails()

                #close program option
                else:
                    closeApp = True
            else:
                print('Please enter a number between 1 - 5 or 77. To EXIT'\
                ' enter 99.')

        else:
            print('Please enter a number from the main menu screen.')
            mainDisplay()

print('>>>Exiting program<<<')

print('Disconnecting from the excel file and closing app.')
print('Thank you for using pyDataViewer.')
wb.close()          #closing the connection to excel workbook
